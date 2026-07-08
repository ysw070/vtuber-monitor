#!/usr/bin/env python3
"""대시보드 HTML 빌더: v8+ xlsx 그룹목록 + collected_vtubers.csv (+스냅샷 히스토리) → dashboard.html
사용: python3 build_dashboard.py <xlsx> <csv> <out_html>
스냅샷 index/diff 는 <csv 폴더>/snapshots/{index.json,diff_latest.json} 에서 자동 로드(있으면)."""
import sys, json, csv, datetime, os
import openpyxl

XLSX, CSVF, OUT = sys.argv[1], sys.argv[2], sys.argv[3]
SNAPDIR = os.path.join(os.path.dirname(os.path.abspath(CSVF)), 'snapshots')

def load_json(p, default):
    try: return json.load(open(p, encoding='utf-8'))
    except Exception: return default
history = load_json(os.path.join(SNAPDIR, 'index.json'), {})
diff = load_json(os.path.join(SNAPDIR, 'diff_latest.json'), {'available': False})
# 유튜브 자동 통계(있으면) — enrich_youtube.py 가 생성, 핸들/채널ID 키
ytstats = load_json(os.path.join(os.path.dirname(os.path.abspath(CSVF)), 'youtube_stats.json'), {})

wb = openpyxl.load_workbook(XLSX)
ws = wb['버추얼 그룹 전수목록']
groups = []
for r in ws.iter_rows(min_row=4, values_only=True):
    if r[0] is None: continue
    groups.append([r[0], r[1] or '', r[2] or '', r[3] or '', str(r[4] or ''), r[5] or '', r[6] or '', r[7] or ''])

# 수동 제보 그룹 병합 (data/manual_groups.csv) — 방문자 제보로 추가된 그룹
MANUAL = os.path.join(os.path.dirname(os.path.abspath(CSVF)), 'manual_groups.csv')
manual_count = 0
if os.path.exists(MANUAL):
    try:
        mrows = list(csv.DictReader(open(MANUAL, encoding='utf-8-sig')))
    except Exception:
        mrows = []
    next_no = max([g[0] for g in groups if isinstance(g[0], int)] or [0]) + 1
    seen = {g[1] for g in groups}
    for m in mrows:
        name = (m.get('name') or '').strip()
        if not name or name in seen:
            continue
        seen.add(name)
        groups.append([
            next_no, name, (m.get('name_en') or '').strip(),
            (m.get('agency') or '').strip() or '(제보)',
            str(m.get('year') or '').strip(),
            (m.get('category') or '').strip() or '스트리밍/MCN형',
            (m.get('status') or '활동중').strip() or '활동중',
            (m.get('note') or '').strip(),
            (m.get('channel') or '').strip(),  # g[8] 채널
            ('검증' if str(m.get('verified') or '').strip().lower() in ('1', 'true', 'y')
             else '제보'),                       # g[9] 출처(검증완료 / 제보·미검증)
        ])
        next_no += 1
        manual_count += 1

rows = list(csv.DictReader(open(CSVF, encoding='utf-8-sig')))
agg = {}
collected_at = ''
for r in rows:
    collected_at = max(collected_at, (r.get('collected_at') or '')[:10])
    n = r['name']
    try: f = int(float(r['followers'] or 0))
    except: f = 0
    if n not in agg:
        agg[n] = {'n': n, 'e': r['name_en'] or '', 'f': 0, 'p': '', 'd': r['debut_date'] or '', 'l': '', 't': r['twitter_id'] or '', 'ch': {}}
    a = agg[n]
    plat = r['platform']
    if plat not in a['ch'] or f > a['ch'][plat][0]:
        a['ch'][plat] = [f, r['channel_url'] or '', r['channel_name'] or '']
    if f > a['f']: a['f'], a['p'] = f, plat
    if r['debut_date'] and (not a['d'] or r['debut_date'] < a['d']): a['d'] = r['debut_date']
    ll = (r['last_live_at'] or '')[:10]
    if ll > a['l']: a['l'] = ll
    if r['twitter_id'] and not a['t']: a['t'] = r['twitter_id']
solos = sorted(agg.values(), key=lambda x: -x['f'])

# 소속사·운영사 정리 시트 → [이름, 유형, 대표그룹, 비고]
agencies = []
if '소속사·운영사 정리' in wb.sheetnames:
    aw = wb['소속사·운영사 정리']
    for r in aw.iter_rows(min_row=3, values_only=True):
        if not r or not r[0]: continue
        agencies.append([r[0] or '', r[1] or '', r[2] or '', (r[3] if len(r) > 3 else '') or ''])

# 데뷔예정·신규 시트 → [그룹, 소속사, 상태, 데뷔(예정)일, 비고]
upcoming = []
for sn in wb.sheetnames:
    if '데뷔예정' in sn or '신규' in sn:
        uw = wb[sn]
        for r in uw.iter_rows(min_row=3, values_only=True):
            if not r or not r[0]: continue
            upcoming.append([r[0] or '', r[1] or '', r[2] or '', str(r[3] or ''), (r[4] if len(r) > 4 else '') or ''])
        break

# 연표: 그룹의 활동시작에서 연도 추출 → 연도별 데뷔/종료 집계 + 항목
import re as _re
def year_of(s):
    m = _re.search(r'(20\d{2})', str(s))
    return m.group(1) if m else None
timeline = {}
for g in groups:
    y = year_of(g[4])
    if not y: continue
    timeline.setdefault(y, {'debut': [], 'ended': []})
    timeline[y]['debut'].append({'name': g[1], 'cls': g[5], 'st': g[6]})
# 종료 연도는 비고/상태에서 추정 어려워 데뷔 기준 연표만 사용(상태로 색 구분)

# 빌드 모드: cowork(기본, 아티팩트용) 또는 public(외부 사이트용)
MODE = os.environ.get('SITE_MODE', 'cowork')
GH_REPO = os.environ.get('GH_REPO', '')  # 예: 'ysw070/vtuber-monitor'
cfg = {'mode': MODE, 'repo': GH_REPO}
# Supabase(회원·권한) — data/supabase_config.json 있으면 연결 (public 모드 전용)
sbconf = load_json(os.path.join(os.path.dirname(os.path.abspath(CSVF)), 'supabase_config.json'), None)
if sbconf and MODE == 'public':
    cfg['supabase'] = {'url': sbconf.get('url', ''), 'key': sbconf.get('key', '')}

data = {'built': datetime.date.today().isoformat(), 'collected': collected_at or datetime.date.today().isoformat(),
        'groups': groups, 'solos': solos, 'history': history, 'diff': diff, 'cfg': cfg,
        'agencies': agencies, 'upcoming': upcoming, 'timeline': timeline, 'ytstats': ytstats}
DATA = json.dumps(data, ensure_ascii=False, separators=(',', ':')).replace('</', '<\\/')

HTML = r'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>한국 버추얼 아이돌·버튜버 모니터링</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.0/dist/chart.umd.js" integrity="sha384-iU8HYtnGQ8Cy4zl7gbNMOhsDTTKX02BTXptVP/vqAWIaTfM7isw76iyZCsjL2eVi" crossorigin="anonymous"></script>
__SUPABASE_SCRIPT__
<style>
:root{color-scheme:light}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif;background:#f6f7fb;color:#1e2433;font-size:14px}
.wrap{max-width:1200px;margin:0 auto;padding:18px}
header{display:flex;flex-wrap:wrap;align-items:center;gap:10px;margin-bottom:14px}
h1{font-size:19px;font-weight:800}
.meta{font-size:12px;color:#7a8194}
.spacer{flex:1}
button{border:0;border-radius:8px;padding:8px 14px;font-size:13px;font-weight:600;cursor:pointer}
.btn-p{background:#3d5afe;color:#fff}.btn-p:hover{background:#2a46e8}
.btn-s{background:#e8ebf4;color:#2c3550}.btn-s:hover{background:#dde1ee}
.tabs{display:flex;gap:6px;margin-bottom:14px}
.tab{padding:8px 16px;border-radius:8px;background:#e8ebf4;color:#555e76;font-weight:700;cursor:pointer;font-size:13px}
.tab.on{background:#1e2433;color:#fff}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:14px}
.card{background:#fff;border-radius:12px;padding:14px;box-shadow:0 1px 3px rgba(20,30,60,.07)}
.card .v{font-size:24px;font-weight:800}.card .k{font-size:12px;color:#7a8194;margin-top:2px}
.charts{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:12px}
.chart-box{background:#fff;border-radius:12px;padding:14px;box-shadow:0 1px 3px rgba(20,30,60,.07)}
.chart-box h3{font-size:13px;margin-bottom:8px;color:#2c3550}
.toolbar{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px;align-items:center}
input[type=text],select{padding:8px 10px;border:1px solid #d7dbe8;border-radius:8px;font-size:13px;background:#fff}
input[type=text]{min-width:200px}
.chip{padding:5px 11px;border-radius:99px;font-size:12px;font-weight:600;cursor:pointer;background:#e8ebf4;color:#555e76}
.chip.on{background:#1e2433;color:#fff}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(20,30,60,.07)}
th{background:#fafbfe;text-align:left;padding:9px 10px;font-size:12px;color:#555e76;cursor:pointer;user-select:none;white-space:nowrap;border-bottom:1px solid #edf0f7}
td{padding:8px 10px;border-bottom:1px solid #f1f3f9;font-size:13px;vertical-align:middle}
tr.row{cursor:pointer}tr.row:hover{background:#f4f6fd}
.st{display:inline-block;padding:2px 9px;border-radius:99px;font-size:11.5px;font-weight:700}
.st-활동중{background:#dcf5e2;color:#157a36}.st-해체종료{background:#fde3e3;color:#b3261e}
.st-휴면{background:#e9e9ec;color:#5d6068}.st-확인필요{background:#dfe9fb;color:#2455c3}.st-전환{background:#fdf0d5;color:#9a6b00}
.num{text-align:right;font-variant-numeric:tabular-nums}
.muted{color:#9aa1b3;font-size:12px}
#drawer{position:fixed;top:0;right:-420px;width:400px;max-width:92vw;height:100%;background:#fff;box-shadow:-6px 0 24px rgba(20,30,60,.15);transition:right .22s;z-index:50;padding:20px;overflow-y:auto}
#drawer.open{right:0}
#drawer h2{font-size:18px;margin-bottom:2px}
#drawer .close{position:absolute;top:14px;right:14px;background:#e8ebf4;border-radius:8px;width:30px;height:30px;font-size:15px}
.kv{margin-top:14px}.kv div{display:flex;gap:10px;padding:6px 0;border-bottom:1px solid #f1f3f9;font-size:13px}
.kv b{min-width:84px;color:#7a8194;font-weight:600}
.chlink{display:flex;justify-content:space-between;align-items:center;background:#f6f7fb;border-radius:10px;padding:10px 12px;margin-top:8px;text-decoration:none;color:#1e2433}
.chlink:hover{background:#eceef7}.chlink .pf{font-weight:700;font-size:13px}.chlink .fl{font-size:12px;color:#7a8194}
.pagei{font-size:12px;color:#7a8194;margin:8px 2px}
.notice{display:none;background:#fff7e0;border:1px solid #f3df9d;color:#7a5b00;border-radius:10px;padding:10px 14px;margin-bottom:12px;font-size:13px}
footer{margin-top:18px;font-size:11.5px;color:#9aa1b3}
#t-submit label{display:block;font-size:12px;color:#555e76;font-weight:600}
#t-submit input,#t-submit select{margin-top:4px}
.st-제보{background:#efe7fb;color:#6b3fc0}
.st-검증{background:#dff3ee;color:#0f7c62}
#auth-modal{display:none;position:fixed;inset:0;background:rgba(20,30,60,.45);z-index:60;align-items:center;justify-content:center}
#auth-modal.open{display:flex}
.am-card{position:relative;background:#fff;border-radius:14px;padding:24px;width:360px;max-width:92vw;box-shadow:0 12px 40px rgba(20,30,60,.25)}
.am-card h2{font-size:18px}
.role-pill{display:inline-block;padding:1px 8px;border-radius:99px;font-size:11px;font-weight:700;background:#e8ebf4;color:#2c3550}
.mrow{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:8px 0;border-bottom:1px solid #f1f3f9;font-size:13px}
@media(max-width:640px){.hide-m{display:none}}
</style></head><body><div class="wrap">
<header>
  <div><h1>한국 버추얼 아이돌·버튜버 모니터링</h1>
  <div class="meta">데이터 수집일 <b id="m-col"></b> · 페이지 빌드 <span id="m-built"></span> · 출처: 전수조사 xlsx + 나의 작은 버튜버 API</div></div>
  <div class="spacer"></div>
  <button class="btn-p" id="btn-collect">🔄 지금 재수집</button>
  <button class="btn-s" id="btn-xg">⬇ 그룹 CSV</button>
  <button class="btn-s" id="btn-xs">⬇ 스트리머 CSV</button>
  <span id="auth-area"></span>
</header>
<div class="notice" id="notice"></div>
<div class="tabs">
  <div class="tab on" data-t="dash">📊 대시보드</div>
  <div class="tab" data-t="grp">그룹 전수목록 (<span id="cnt-g"></span>)</div>
  <div class="tab" data-t="solo">수집 스트리머 (<span id="cnt-s"></span>)</div>
  <div class="tab" data-t="agency">소속사 (<span id="cnt-a"></span>)</div>
  <div class="tab" data-t="upcoming">데뷔예정 (<span id="cnt-u"></span>)</div>
  <div class="tab" data-t="timeline">📅 연표</div>
  <div class="tab" data-t="trend">📈 추세·변동</div>
  <div class="tab" data-t="method">ℹ️ 데이터 기준</div>
  <div class="tab" data-t="premium">⭐ 프리미엄</div>
  <div class="tab" data-t="submit">➕ 그룹 제보</div>
  <div class="tab" data-t="admin">🛠 관리자</div>
</div>

<section id="t-dash">
  <div class="cards" id="cards"></div>
  <div class="charts">
    <div class="chart-box"><h3>그룹 상태별</h3><canvas id="c1"></canvas></div>
    <div class="chart-box"><h3>그룹 분류별</h3><canvas id="c2"></canvas></div>
    <div class="chart-box"><h3>스트리머 규모 분포 (최대 팔로워)</h3><canvas id="c3"></canvas></div>
    <div class="chart-box"><h3>스트리머 주 플랫폼</h3><canvas id="c4"></canvas></div>
  </div>
</section>

<section id="t-grp" style="display:none">
  <div class="toolbar">
    <input type="text" id="q-g" placeholder="그룹명·소속사 검색">
    <span class="chip on" data-s="">전체</span><span class="chip" data-s="활동중">활동중</span><span class="chip" data-s="해체/종료">해체/종료</span><span class="chip" data-s="휴면">휴면</span><span class="chip" data-s="확인필요">확인필요</span>
  </div>
  <table><thead><tr><th data-k="0">No.</th><th data-k="1">그룹명</th><th data-k="3">소속사·운영사</th><th data-k="4" class="hide-m">시작</th><th data-k="5" class="hide-m">분류</th><th data-k="6">상태</th></tr></thead><tbody id="tb-g"></tbody></table>
</section>

<section id="t-solo" style="display:none">
  <p class="muted" style="margin-bottom:10px;line-height:1.55">라이브 스트리머 집계 API로 자동 수집한 목록이에요. <b>이 소스는 소속사 정보를 제공하지 않아 소속 여부를 판별할 수 없습니다</b>(소속 미상). 따라서 '무소속(개인세)' 목록이 아니며, 소속 있는 스트리머가 포함될 수 있어요. 소속이 확인된 솔로 아티스트는 <b>그룹 전수목록</b>의 <code>버추얼휴먼/솔로</code>로 등재합니다.</p>
  <div class="toolbar">
    <input type="text" id="q-s" placeholder="활동명 검색">
    <select id="f-plat"><option value="">전체 플랫폼</option></select>
    <select id="f-min"><option value="0">규모 전체</option><option value="10000">1만+</option><option value="30000">3만+</option><option value="50000">5만+ (포함 후보)</option></select>
  </div>
  <table><thead><tr><th data-k="n">활동명</th><th data-k="f" class="num">최대 팔로워</th><th data-k="p" class="hide-m">주 플랫폼</th><th data-k="d" class="hide-m">데뷔일</th><th data-k="l">최근 라이브</th></tr></thead><tbody id="tb-s"></tbody></table>
  <div class="pagei" id="pg-s"></div>
</section>

<section id="t-trend" style="display:none">
  <div id="trend-empty" class="notice" style="display:none;background:#eef1fb;border-color:#cfd8f5;color:#3a4a7a"></div>
  <div id="trend-body" style="display:none">
    <div class="charts">
      <div class="chart-box"><h3>수집 인원 추이</h3><canvas id="ct1"></canvas></div>
      <div class="chart-box"><h3>규모 구간별 인원 추이</h3><canvas id="ct2"></canvas></div>
    </div>
    <div id="diff-wrap" style="margin-top:14px"></div>
  </div>
</section>

<section id="t-agency" style="display:none">
  <div class="toolbar"><input type="text" id="q-a" placeholder="소속사·대표그룹 검색"></div>
  <table><thead><tr><th data-k="0">소속사·운영사</th><th data-k="1" class="hide-m">유형</th><th data-k="2">대표 그룹</th></tr></thead><tbody id="tb-a"></tbody></table>
</section>

<section id="t-upcoming" style="display:none">
  <p class="muted" style="margin-bottom:10px">2026 신규 데뷔 및 데뷔 예정 워치리스트 — 정식 데뷔 시 그룹 전수목록으로 편입됩니다.</p>
  <div id="up-cards" class="charts"></div>
</section>

<section id="t-timeline" style="display:none">
  <p class="muted" style="margin-bottom:10px">데뷔(활동시작) 연도 기준. 색은 현재 상태 — 한국 버추얼 그룹 신(scene)의 흐름을 한눈에.</p>
  <div class="chart-box" style="margin-bottom:14px"><h3>연도별 데뷔 수</h3><canvas id="ctl"></canvas></div>
  <div id="tl-body"></div>
</section>

<section id="t-method" style="display:none"><div id="method-body"></div></section>

<section id="t-premium" style="display:none"><div id="premium-body"></div></section>

<section id="t-submit" style="display:none">
  <div class="chart-box" style="max-width:640px">
    <h3 style="font-size:16px;margin-bottom:4px">➕ 빠진 그룹 제보하기</h3>
    <p class="muted" style="line-height:1.55;margin-bottom:14px">위키 미등재·조용히 활동하는 그룹이 빠졌다면 알려주세요. 제출하면 제보 게시판에 접수되고, <b>관리자 승인 후</b> 명부에 <b>'제보' 표시</b>와 함께 추가됩니다.</p>
    <div style="display:grid;gap:10px">
      <label>그룹명 <span style="color:#e74c3c">*</span><input type="text" id="gf-name" style="width:100%" placeholder="예: 별빛소녀단"></label>
      <label>영문/별칭<input type="text" id="gf-en" style="width:100%" placeholder="예: Starlight Girls"></label>
      <label>소속사·운영사<input type="text" id="gf-agency" style="width:100%" placeholder="모르면 비워두세요"></label>
      <div style="display:flex;gap:10px;flex-wrap:wrap">
        <label style="flex:1;min-width:120px">활동시작(연도)<input type="text" id="gf-year" style="width:100%" placeholder="예: 2025"></label>
        <label style="flex:1;min-width:140px">분류<select id="gf-cat" style="width:100%"><option>스트리밍/MCN형</option><option>음반/아이돌형</option><option>버추얼휴먼/솔로</option><option>혼합형</option></select></label>
        <label style="flex:1;min-width:120px">상태<select id="gf-status" style="width:100%"><option>활동중</option><option>휴면</option><option>확인필요</option></select></label>
      </div>
      <label>대표 채널 링크 (유튜브/치지직/SOOP/X)<input type="text" id="gf-channel" style="width:100%" placeholder="https://..."></label>
      <label>비고<textarea id="gf-note" rows="3" style="width:100%;padding:8px 10px;border:1px solid #d7dbe8;border-radius:8px;font-size:13px;font-family:inherit" placeholder="멤버 수, 데뷔곡, 플랫폼 등 아는 정보"></textarea></label>
      <label>제보자(선택)<input type="text" id="gf-submitter" style="width:100%" placeholder="닉네임 또는 빈칸"></label>
      <button class="btn-p" id="gf-submit" style="margin-top:4px">제보 보내기 (GitHub 이슈 열기)</button>
      <div id="gf-msg" class="muted"></div>
    </div>
  </div>
</section>

<section id="t-admin" style="display:none">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">
    <h3 style="font-size:16px">🛠 제보 게시판 (관리자)</h3>
    <button class="btn-s" id="admin-refresh" style="padding:5px 10px">새로고침</button>
  </div>
  <p class="muted" style="line-height:1.55;margin-bottom:12px">에디터·관리자 전용 페이지예요. 검수 대기 제보를 <b>승인</b>하면 즉시 명부에 반영되고, <b>반려</b>하면 숨겨집니다. 관리자는 아래에서 회원 등급(일반/에디터/관리자)도 바꿀 수 있어요.</p>
  <div id="admin-body"></div>
</section>

<div id="auth-modal"><div class="am-card">
  <button class="close" onclick="closeAuthModal()">✕</button>
  <h2 id="am-title">로그인</h2>
  <p class="muted" style="margin:4px 0 12px">회원가입하면 그룹 제보를 할 수 있어요. 가입 직후엔 일반회원 등급입니다.</p>
  <label style="display:block;font-size:12px;color:#555e76;font-weight:600">이메일<input type="email" id="am-email" style="width:100%;margin-top:4px" placeholder="you@example.com"></label>
  <label style="display:block;font-size:12px;color:#555e76;font-weight:600;margin-top:10px">비밀번호<input type="password" id="am-pw" style="width:100%;margin-top:4px" placeholder="6자 이상"></label>
  <div style="display:flex;gap:8px;margin-top:14px">
    <button class="btn-p" id="am-login" style="flex:1">로그인</button>
    <button class="btn-s" id="am-signup" style="flex:1">회원가입</button>
  </div>
  <div id="am-msg" class="muted" style="margin-top:8px;min-height:16px"></div>
</div></div>

<div id="drawer"><button class="close" onclick="drawer.classList.remove('open')">✕</button><div id="d-body"></div></div>
<footer id="foot">임계값(확정): 위키등재 OR 5만+ 팔로워 · 매월 1일 자동 수집.</footer>
<div id="copyright" style="margin-top:8px;font-size:11.5px;color:#9aa1b3">© <span id="cyear"></span> (주)크리에이터버스. All rights reserved.</div>
</div>
<script>
const D=__DATA__;
const TASK='__TASKID__';
const $=s=>document.querySelector(s),esc=s=>String(s??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
const fmt=n=>n>=10000?(n/10000).toFixed(n>=100000?0:1)+'만':n.toLocaleString();
$('#m-col').textContent=D.collected;$('#m-built').textContent=D.built;
$('#cnt-g').textContent=D.groups.length;$('#cnt-s').textContent=D.solos.length;
$('#cnt-a').textContent=(D.agencies||[]).length;$('#cnt-u').textContent=(D.upcoming||[]).length;
const drawer=$('#drawer');
const TABS=['dash','grp','solo','agency','upcoming','timeline','trend','method','premium','submit','admin'];

/* 탭 */
function showTab(t){
  document.querySelectorAll('.tab').forEach(x=>x.classList.toggle('on',x.dataset.t===t));
  TABS.forEach(k=>$('#t-'+k).style.display=k===t?'':'none');
  if(t==='admin') renderAdmin();
  if(t==='premium') renderPremium();
}
document.querySelectorAll('.tab').forEach(t=>t.onclick=()=>showTab(t.dataset.t));

/* 대시보드 */
const cnt=(arr,fn)=>arr.reduce((m,x)=>{const k=fn(x);m[k]=(m[k]||0)+1;return m},{});
const gs=cnt(D.groups,g=>g[6]),gc=cnt(D.groups,g=>g[5]);
const n5=D.solos.filter(s=>s.f>=50000).length,n3=D.solos.filter(s=>s.f>=30000).length;
$('#cards').innerHTML=[['전체 그룹',D.groups.length],['활동중 그룹',gs['활동중']||0],['수집 스트리머',D.solos.length],['5만+ 포함 후보',n5],['3만+ 워치',n3-n5]]
 .map(([k,v])=>`<div class="card"><div class="v">${v}</div><div class="k">${k}</div></div>`).join('');
const PAL=['#3d5afe','#27ae60','#e74c3c','#95a0b6','#f5a623','#8e44ad','#16a2b8'];
const mk=(id,type,labels,vals,horiz)=>new Chart($(id),{type,data:{labels,datasets:[{data:vals,backgroundColor:PAL,borderWidth:0}]},
 options:{indexAxis:horiz?'y':'x',plugins:{legend:{display:type==='doughnut',position:'right',labels:{boxWidth:12,font:{size:11}}}},
 scales:type==='doughnut'?{}:{x:{ticks:{font:{size:11}}},y:{ticks:{font:{size:11}}}}}});
mk('#c1','doughnut',Object.keys(gs),Object.values(gs));
mk('#c2','bar',Object.keys(gc),Object.values(gc),true);
const bks=[['5만+',s=>s.f>=50000],['3만~5만',s=>s.f>=30000&&s.f<50000],['1만~3만',s=>s.f>=10000&&s.f<30000],['1천~1만',s=>s.f>=1000&&s.f<10000],['1천 미만',s=>s.f<1000]];
mk('#c3','bar',bks.map(b=>b[0]),bks.map(b=>D.solos.filter(b[1]).length));
const sp=cnt(D.solos,s=>s.p||'미상');
mk('#c4','doughnut',Object.keys(sp),Object.values(sp));

/* 유튜브 자동 통계 조회: 채널 URL → 캐시 키(@핸들/UC아이디) */
function ytLookup(url){
  if(!url||!D.ytstats)return null;
  let m=url.match(/youtube\.com\/@([A-Za-z0-9._\-가-힣]+)/);
  if(m&&D.ytstats['@'+m[1]])return D.ytstats['@'+m[1]];
  m=url.match(/youtube\.com\/channel\/(UC[A-Za-z0-9_\-]+)/);
  if(m&&D.ytstats[m[1]])return D.ytstats[m[1]];
  return null;
}

/* 그룹 테이블 */
let gQ='',gS='',gSort=[0,1];
const stCls=s=>'st st-'+s.replace('/','');
function renderG(){
  let rs=D.groups.filter(g=>(!gS||g[6]===gS)&&(!gQ||(g[1]+g[2]+g[3]).toLowerCase().includes(gQ)));
  rs.sort((a,b)=>{const k=gSort[0];return (a[k]>b[k]?1:a[k]<b[k]?-1:0)*gSort[1]});
  $('#tb-g').innerHTML=rs.map((g,i)=>`<tr class="row" data-i="${D.groups.indexOf(g)}"><td class="muted">${g[0]}</td><td><b>${esc(g[1])}</b> <span class="muted">${esc(g[2])}</span>${g[9]==='검증'?' <span class="st st-검증">검증</span>':(g[9]==='제보'?' <span class="st st-제보">제보</span>':'')}</td><td>${esc(g[3])}</td><td class="hide-m">${esc(g[4])}</td><td class="hide-m muted">${esc(g[5])}</td><td><span class="${stCls(g[6])}">${esc(g[6])}</span></td></tr>`).join('');
  document.querySelectorAll('#tb-g .row').forEach(tr=>tr.onclick=()=>openG(D.groups[+tr.dataset.i]));
}
$('#q-g').oninput=e=>{gQ=e.target.value.toLowerCase();renderG()};
document.querySelectorAll('#t-grp .chip').forEach(c=>c.onclick=()=>{document.querySelectorAll('#t-grp .chip').forEach(x=>x.classList.remove('on'));c.classList.add('on');gS=c.dataset.s;renderG()});
document.querySelectorAll('#t-grp th').forEach(th=>th.onclick=()=>{const k=+th.dataset.k;gSort=[k,gSort[0]===k?-gSort[1]:1];renderG()});
function openG(g,skipHash){
  if(!skipHash)setHash('g/'+g[0]);
  const isSub=g[9]==='제보', isVer=g[9]==='검증';
  const subBadge=isVer?' <span class="st st-검증">검증완료</span>':(isSub?' <span class="st st-제보">제보·미검증</span>':'');
  const chLink=g[8]?`<a class="chlink" target="_blank" href="${esc(g[8])}"><span class="pf">🔗 공식 채널</span><span class="fl">새 창 ↗</span></a>`:'';
  const yt=ytLookup(g[8]);
  const ytLine=yt?`<div class="kv"><div><b>📊 유튜브</b><span>구독 <b>${yt.subscribers!=null?fmt(yt.subscribers):'비공개'}</b>${yt.videos!=null?' · 영상 '+yt.videos.toLocaleString()+'개':''}${yt.views?' · 누적 '+fmt(yt.views)+'뷰':''} <span class="muted">(자동 갱신 ${esc(yt.fetched||'')})</span></span></div></div>`:'';
  $('#d-body').innerHTML=`<h2>${esc(g[1])}${subBadge}</h2><div class="muted">${esc(g[2])}</div>
  <div class="kv"><div><b>상태</b><span class="${stCls(g[6])}">${esc(g[6])}</span></div><div><b>소속·운영사</b><span>${esc(g[3])||'—'}</span></div>
  <div><b>활동 시작</b><span>${esc(g[4])||'—'}</span></div><div><b>분류</b><span>${esc(g[5])}</span></div></div>
  <div class="kv"><div style="display:block"><b>비고</b><p style="margin-top:6px;line-height:1.55">${esc(g[7])||'—'}</p></div></div>
  ${ytLine}
  ${chLink}
  <a class="chlink" target="_blank" href="https://www.youtube.com/results?search_query=${encodeURIComponent(g[1]+' 버추얼')}"><span class="pf">▶ 유튜브에서 검색</span><span class="fl">새 창</span></a>
  <a class="chlink" target="_blank" href="https://namu.wiki/Search?q=${encodeURIComponent(g[1])}"><span class="pf">📖 나무위키 검색</span><span class="fl">새 창</span></a>
  ${isSub?'<p class="muted" style="margin-top:10px">※ 방문자 제보로 추가된 항목입니다. 운영자 검증 전이에요.</p>':''}
  ${isVer?'<p class="muted" style="margin-top:10px">※ 명부에 추가 등재된 항목으로, 운영자가 공식 채널·언론·위키로 교차검증했습니다.</p>':''}
  ${shareBtn()}`;
  drawer.classList.add('open');
}

/* 개인세 테이블 */
let sQ='',sP='',sM=0,sSort=['f',-1],sLimit=100;
[...new Set(D.solos.map(s=>s.p).filter(Boolean))].forEach(p=>$('#f-plat').insertAdjacentHTML('beforeend',`<option>${esc(p)}</option>`));
function renderS(){
  let rs=D.solos.filter(s=>s.f>=sM&&(!sP||Object.keys(s.ch).includes(sP))&&(!sQ||(s.n+s.e).toLowerCase().includes(sQ)));
  rs.sort((a,b)=>{const k=sSort[0];return (a[k]>b[k]?1:a[k]<b[k]?-1:0)*sSort[1]});
  const total=rs.length;rs=rs.slice(0,sLimit);
  $('#tb-s').innerHTML=rs.map(s=>`<tr class="row" data-n="${esc(s.n)}"><td><b>${esc(s.n)}</b> <span class="muted">${esc(s.e)}</span></td><td class="num"><b>${s.f?fmt(s.f):'—'}</b></td><td class="hide-m">${esc(s.p)||'—'}</td><td class="hide-m muted">${esc(s.d)||'—'}</td><td class="muted">${esc(s.l)||'—'}</td></tr>`).join('');
  $('#pg-s').innerHTML=total>sLimit?`${total}건 중 ${sLimit}건 표시 — <a href="#" id="more">더 보기 (+100)</a>`:`${total}건 표시`;
  const m=$('#more');if(m)m.onclick=e=>{e.preventDefault();sLimit+=100;renderS()};
  document.querySelectorAll('#tb-s .row').forEach(tr=>tr.onclick=()=>openS(D.solos.find(x=>x.n===tr.dataset.n)));
}
$('#q-s').oninput=e=>{sQ=e.target.value.toLowerCase();sLimit=100;renderS()};
$('#f-plat').onchange=e=>{sP=e.target.value;sLimit=100;renderS()};
$('#f-min').onchange=e=>{sM=+e.target.value;sLimit=100;renderS()};
document.querySelectorAll('#t-solo th').forEach(th=>th.onclick=()=>{const k=th.dataset.k;sSort=[k,sSort[0]===k?-sSort[1]:-1];renderS()});
function openS(s,skipHash){
  if(!skipHash)setHash('s/'+encodeURIComponent(s.n));
  const links=Object.entries(s.ch).map(([p,[f,u,cn]])=>u?`<a class="chlink" target="_blank" href="${esc(u)}"><span class="pf">${esc(p)} · ${esc(cn||s.n)}</span><span class="fl">${f?fmt(f)+' 팔로워':''} ↗</span></a>`:'').join('');
  const tw=s.t?`<a class="chlink" target="_blank" href="https://x.com/${esc(s.t.replace('@',''))}"><span class="pf">𝕏 ${esc(s.t)}</span><span class="fl">새 창 ↗</span></a>`:'';
  const judge=s.f>=50000?'<span class="st st-활동중">5만+ 포함 후보</span>':s.f>=30000?'<span class="st st-전환">3만~5만 워치</span>':'<span class="st st-휴면">기준 미달 (위키등재 별도확인)</span>';
  $('#d-body').innerHTML=`<h2>${esc(s.n)}</h2><div class="muted">${esc(s.e)}</div>
  <div class="kv"><div><b>임계값 판정</b><span>${judge}</span></div><div><b>최대 팔로워</b><span><b>${s.f?s.f.toLocaleString():'—'}</b> (${esc(s.p)||'—'})</span></div>
  <div><b>데뷔일</b><span>${esc(s.d)||'—'}</span></div><div><b>최근 라이브</b><span>${esc(s.l)||'—'}</span></div></div>
  <h3 style="margin-top:16px;font-size:13px;color:#555e76">채널</h3>${links||'<p class="muted" style="margin-top:6px">채널 URL 없음</p>'}${tw}
  ${shareBtn()}`;
  drawer.classList.add('open');
}

/* CSV 다운로드 (BOM 포함 → 엑셀에서 바로 열림) */
function dl(name,rows){
  const csv='﻿'+rows.map(r=>r.map(c=>`"${String(c??'').replace(/"/g,'""')}"`).join(',')).join('\r\n');
  const a=document.createElement('a');a.href=URL.createObjectURL(new Blob([csv],{type:'text/csv'}));a.download=name;a.click();
}
$('#btn-xg').onclick=()=>dl(`버추얼그룹_전수목록_${D.collected}.csv`,[['No.','그룹명','영문/별칭','소속사·운영사','활동시작','분류','상태','비고'],...D.groups]);
$('#btn-xs').onclick=()=>dl(`수집스트리머_${D.collected}.csv`,[['활동명','영문','최대 팔로워','주 플랫폼','플랫폼 목록','데뷔일','최근 라이브','트위터'],...D.solos.map(s=>[s.n,s.e,s.f,s.p,Object.keys(s.ch).join('·'),s.d,s.l,s.t])]);

/* 재수집 버튼 — 모드별 분기 */
const CFG=D.cfg||{mode:'cowork'};
const btn=$('#btn-collect'),notice=$('#notice');
if(CFG.mode==='public'){
  // 공개 사이트: 시크릿 노출 없이 GitHub Actions 'Run workflow' 페이지로 이동 (레포 권한자만 실제 실행 가능)
  btn.textContent='🔄 수동 수집 실행';
  const url=CFG.repo?`https://github.com/${CFG.repo}/actions/workflows/monthly-collect.yml`:'';
  btn.onclick=()=>{
    if(url)window.open(url,'_blank');
    notice.style.display='block';
    notice.innerHTML=url?`GitHub Actions 페이지를 열었어요. <b>Run workflow</b> 버튼을 누르면 수집→스냅샷→재배포가 클라우드에서 실행됩니다(레포 권한 필요, 수 분 후 사이트 자동 갱신).`:'레포가 설정되지 않았습니다. README의 GH_REPO 안내를 확인하세요.';
  };
}else{
  // Cowork 아티팩트: 월간 스케줄 작업 즉시 실행
  btn.onclick=async()=>{
    try{
      if(!window.cowork||!window.cowork.runScheduledTask)throw 0;
      await window.cowork.runScheduledTask(TASK);
      notice.style.display='block';notice.textContent='수집 작업을 시작했어요. 완료까지 수 분 걸리며, 끝나면 이 페이지가 갱신됩니다 — 잠시 후 다시 열어 확인해 주세요.';
    }catch(e){
      notice.style.display='block';notice.textContent='작업 실행에 실패했어요. 채팅에서 "버튜버 데이터 재수집해줘"라고 요청해 주세요.';
    }
  };
}

/* 추세·변동 */
function renderTrend(){
  const hist=D.history||{},keys=Object.keys(hist).sort();
  const empty=$('#trend-empty'),body=$('#trend-body');
  if(keys.length<2){
    empty.style.display='block';body.style.display='none';
    empty.innerHTML=`스냅샷이 <b>${keys.length}개</b> 있어요. 추세 차트와 변동 리포트는 <b>스냅샷이 2개 이상</b> 쌓이면 나타납니다 — 다음 월간 수집(또는 "지금 재수집") 이후 자동으로 채워져요.${keys.length?`<br><span class="muted">현재 스냅샷: ${keys.join(', ')} · 인원 ${hist[keys[0]].total_people}명</span>`:''}`;
    return;
  }
  empty.style.display='none';body.style.display='block';
  new Chart($('#ct1'),{type:'line',data:{labels:keys,datasets:[
    {label:'총 인원',data:keys.map(k=>hist[k].total_people),borderColor:'#3d5afe',backgroundColor:'#3d5afe22',tension:.3,fill:true},
    {label:'총 채널',data:keys.map(k=>hist[k].total_channels),borderColor:'#16a2b8',backgroundColor:'#16a2b822',tension:.3,fill:true}
  ]},options:{plugins:{legend:{labels:{font:{size:11}}}},scales:{y:{beginAtZero:false,ticks:{font:{size:11}}},x:{ticks:{font:{size:11}}}}}});
  new Chart($('#ct2'),{type:'line',data:{labels:keys,datasets:[
    {label:'5만+',data:keys.map(k=>hist[k].n_50k),borderColor:'#e74c3c',tension:.3},
    {label:'3만+',data:keys.map(k=>hist[k].n_30k),borderColor:'#f5a623',tension:.3},
    {label:'1만+',data:keys.map(k=>hist[k].n_10k),borderColor:'#27ae60',tension:.3}
  ]},options:{plugins:{legend:{labels:{font:{size:11}}}},scales:{y:{beginAtZero:true,ticks:{font:{size:11}}},x:{ticks:{font:{size:11}}}}}});
  const df=D.diff||{};
  if(df.available){
    const lst=(t,arr,fn)=>`<div class="chart-box" style="margin-bottom:12px"><h3>${t} (${arr.length})</h3>${arr.length?'<div style="display:flex;flex-wrap:wrap;gap:6px">'+arr.slice(0,40).map(fn).join('')+'</div>':'<span class="muted">없음</span>'}</div>`;
    const pill=(txt,cls)=>`<span class="st ${cls||'st-확인필요'}">${esc(txt)}</span>`;
    $('#diff-wrap').innerHTML=`<h3 style="font-size:14px;margin-bottom:8px">📋 ${df.from} → ${df.to} 변동 리포트</h3>`+
      lst('🆕 신규 수집 진입',df.entered,n=>pill(n,'st-활동중'))+
      lst('🚪 이탈(목록에서 사라짐)',df.left,n=>pill(n,'st-해체종료'))+
      lst('⬆ 5만+ 신규 진입',df.up50,o=>pill(`${o.name} (${fmt(o.to)})`,'st-활동중'))+
      lst('⬇ 5만 미만으로 하락',df.down50,o=>pill(`${o.name} (${fmt(o.to)})`,'st-휴면'))+
      `<div class="charts"><div class="chart-box"><h3>📈 팔로워 증가 톱</h3>${df.top_gain.filter(m=>m.delta>0).slice(0,10).map(m=>`<div style="display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #f1f3f9;font-size:13px"><span>${esc(m.name)}</span><span class="num" style="color:#157a36;font-weight:700">+${fmt(m.delta)}</span></div>`).join('')||'<span class="muted">없음</span>'}</div>`+
      `<div class="chart-box"><h3>📉 팔로워 감소 톱</h3>${df.top_drop.filter(m=>m.delta<0).slice(0,10).map(m=>`<div style="display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #f1f3f9;font-size:13px"><span>${esc(m.name)}</span><span class="num" style="color:#b3261e;font-weight:700">${fmt(m.delta)}</span></div>`).join('')||'<span class="muted">없음</span>'}</div></div>`;
  }else $('#diff-wrap').innerHTML='';
}

$('#foot').innerHTML=CFG.mode==='public'
  ? '임계값(확정): 위키등재 OR 5만+ 팔로워 · 매월 1일 GitHub Actions에서 자동 수집·재배포 (서버리스, 운영자 PC와 무관). 데이터 출처: 한국 버추얼아이돌 전수조사 + 나의 작은 버튜버 API.'
  : '임계값(확정): 위키등재 OR 5만+ 팔로워 · 재수집 버튼은 월간 수집 작업을 즉시 실행합니다 (완료까지 수 분, 완료 후 페이지 자동 갱신). 매월 1일 09:00 자동 수집.';

/* ── 공유 딥링크 ── */
function setHash(h){ if(location.hash.slice(1)!==h) history.replaceState(null,'','#'+h); }
function shareBtn(){ return `<button class="btn-s" style="margin-top:16px;width:100%" onclick="copyShare()">🔗 이 항목 링크 복사</button><div id="share-msg" class="muted" style="margin-top:6px"></div>`; }
function copyShare(){
  const url=location.href;
  const done=()=>{const m=$('#share-msg');if(m){m.textContent='링크가 복사됐어요!';setTimeout(()=>{if(m)m.textContent='';},2500);}};
  if(navigator.clipboard&&navigator.clipboard.writeText){navigator.clipboard.writeText(url).then(done,()=>prompt('아래 링크를 복사하세요:',url));}
  else prompt('아래 링크를 복사하세요:',url);
}
drawer.addEventListener('transitionend',()=>{ if(!drawer.classList.contains('open')&&location.hash) setHash(''); });
function routeHash(){
  const h=decodeURIComponent(location.hash.slice(1));
  if(!h) return;
  const [kind,...rest]=h.split('/'); const id=rest.join('/');
  if(kind==='g'){const g=D.groups.find(x=>String(x[0])===id); if(g){showTab('grp');openG(g,true);}}
  else if(kind==='s'){const s=D.solos.find(x=>x.n===id); if(s){showTab('solo');openS(s,true);}}
  else if(kind==='a'){const a=(D.agencies||[]).find(x=>x[0]===id); if(a){showTab('agency');openA(a,true);}}
  else if(TABS.includes(kind)){showTab(kind);}
}

/* ── 소속사 탭 ── */
let aQ='';
function renderA(){
  const rs=(D.agencies||[]).filter(a=>!aQ||(a[0]+a[1]+a[2]+a[3]).toLowerCase().includes(aQ));
  $('#tb-a').innerHTML=rs.map(a=>`<tr class="row" data-n="${esc(a[0])}"><td><b>${esc(a[0])}</b></td><td class="hide-m muted">${esc(a[1])}</td><td>${esc(a[2])||'—'}</td></tr>`).join('');
  document.querySelectorAll('#tb-a .row').forEach(tr=>tr.onclick=()=>openA((D.agencies||[]).find(x=>x[0]===tr.dataset.n)));
}
const qa=$('#q-a'); if(qa) qa.oninput=e=>{aQ=e.target.value.toLowerCase();renderA();};
function openA(a,skipHash){
  if(!skipHash)setHash('a/'+encodeURIComponent(a[0]));
  // 이 소속사에 매핑되는 그룹 전수목록 항목 찾기(소속사 컬럼 부분일치)
  const key=a[0].replace(/\(.*?\)/g,'').trim();
  const linked=D.groups.filter(g=>g[3]&&key&&g[3].includes(key.split(/[ (]/)[0]));
  $('#d-body').innerHTML=`<h2>${esc(a[0])}</h2><div class="muted">${esc(a[1])}</div>
  <div class="kv"><div><b>대표 그룹</b><span>${esc(a[2])||'—'}</span></div></div>
  <div class="kv"><div style="display:block"><b>비고</b><p style="margin-top:6px;line-height:1.55">${esc(a[3])||'—'}</p></div></div>
  ${linked.length?`<h3 style="margin-top:16px;font-size:13px;color:#555e76">전수목록 매칭 (${linked.length})</h3>`+linked.map(g=>`<div class="chlink" style="cursor:pointer" onclick="showTab('grp');openG(D.groups[${D.groups.indexOf(g)}])"><span class="pf">${esc(g[1])}</span><span class="fl"><span class="${stCls(g[6])}">${esc(g[6])}</span></span></div>`).join(''):''}
  ${shareBtn()}`;
  drawer.classList.add('open');
}

/* ── 데뷔예정 탭 ── */
function renderU(){
  $('#up-cards').innerHTML=(D.upcoming||[]).map(u=>`<div class="chart-box"><div style="display:flex;justify-content:space-between;align-items:start;gap:8px"><h3 style="font-size:15px">${esc(u[0])}</h3><span class="st st-확인필요">${esc(u[2])}</span></div>
    <div class="kv"><div><b>소속사</b><span>${esc(u[1])||'—'}</span></div><div><b>데뷔(예정)</b><span>${esc(u[3])||'—'}</span></div></div>
    <p class="muted" style="margin-top:8px;line-height:1.5">${esc(u[4])||''}</p></div>`).join('')||'<p class="muted">데뷔예정 데이터 없음</p>';
}

/* ── 연표 탭 ── */
function renderTL(){
  const tl=D.timeline||{},years=Object.keys(tl).sort();
  if(!years.length){$('#tl-body').innerHTML='<p class="muted">연도 데이터 없음</p>';return;}
  new Chart($('#ctl'),{type:'bar',data:{labels:years,datasets:[{label:'데뷔',data:years.map(y=>tl[y].debut.length),backgroundColor:'#3d5afe',borderWidth:0}]},
    options:{plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,ticks:{font:{size:11},precision:0}},x:{ticks:{font:{size:11}}}}}});
  $('#tl-body').innerHTML=years.slice().reverse().map(y=>`<div class="chart-box" style="margin-bottom:10px"><h3>${y} <span class="muted">· ${tl[y].debut.length}팀 데뷔</span></h3>
    <div style="display:flex;flex-wrap:wrap;gap:6px;margin-top:6px">${tl[y].debut.map(d=>`<span class="${stCls(d.st)}" style="cursor:pointer" onclick="openByName('${esc(d.name).replace(/'/g,"\\'")}')">${esc(d.name)}</span>`).join('')}</div></div>`).join('');
}
function openByName(n){const g=D.groups.find(x=>x[1]===n);if(g){showTab('grp');openG(g);}}

/* ── 데이터 기준(방법론) 탭 ── */
function renderMethod(){
  const wikiN=null;
  $('#method-body').innerHTML=`
  <div class="chart-box" style="line-height:1.65">
    <h3 style="font-size:16px;margin-bottom:8px">이 명부는 어떻게 만들어졌나</h3>
    <p>한국에서 활동하는 버추얼 아이돌·버추얼 유튜버(버튜버)를 <b>그룹·개별 엔티티 단위</b>로 전수조사하고, 라이브 스트리머는 <b>규모·등재 기준</b>으로 선별해 함께 수록합니다. 매월 1일 자동으로 갱신됩니다.</p>
    <p style="margin-top:8px">수록 경로는 둘입니다. ① <b>명부(전수목록)</b> — 운영자가 공식 채널·언론·위키로 교차검증해 등재(소속사·분류 포함). ② <b>수집 스트리머</b> — 스트리머 집계 API로 자동 수집(규모 기준 선별).</p>
  </div>

  <div class="chart-box" style="margin-top:12px">
    <h3 style="font-size:15px">① 수집 출처</h3>
    <div class="kv">
      <div><b>그룹 명부</b><span>thewiki.kr·나무위키 '버추얼 그룹 목록', 언론 보도(머니투데이·이데일리·한경·스포츠경향·전자신문·톱셀럽·KMJ 등), MBC 버추얼 라이브 페스티벌 라인업을 교차검증</span></div>
      <div><b>스트리머 규모</b><span>'나의 작은 버튜버' 공개 API(유튜브 구독·치지직/SOOP 팔로워·최근 방송 집계). <b>소속사 정보는 제공하지 않음</b> → 소속 미상</span></div>
      <div><b>채널 통계</b><span>YouTube Data API v3 — 명부 등재 엔티티의 채널 구독자·영상수·조회수를 월간 자동 갱신</span></div>
      <div><b>소속사 확정</b><span>공식 보도자료·공식 채널·위키 문서 (2026-06-06 재조사 반영)</span></div>
    </div>
  </div>

  <div class="chart-box" style="margin-top:12px">
    <h3 style="font-size:15px">② 5기준 분류 체계 (A~E)</h3>
    <table style="margin-top:8px;box-shadow:none">
      <thead><tr><th>축</th><th>정의</th><th class="hide-m">데이터원</th></tr></thead>
      <tbody>
        <tr><td><b>A 활동기간</b></td><td>데뷔(전향)일 경과 — 2년+ / 3년+</td><td class="hide-m muted">위키·공식 SNS 첫 방송</td></tr>
        <tr><td><b>B 규모</b></td><td>유튜브 구독 또는 치지직/SOOP 팔로워 — 1만 / 3만 / 5만 단계</td><td class="hide-m muted">YouTube·치지직·SOOP API</td></tr>
        <tr><td><b>C 등재성</b></td><td>나무위키·더위키 개별 문서 존재</td><td class="hide-m muted">위키 검색</td></tr>
        <tr><td><b>D 공적활동</b></td><td>음원·페스티벌·공중파·언론 보도 1건+</td><td class="hide-m muted">음원사·뉴스</td></tr>
        <tr><td><b>E 인증</b></td><td>플랫폼 공식 인증(치지직 파트너·SOOP 파트너BJ·YT)</td><td class="hide-m muted">각 플랫폼</td></tr>
      </tbody>
    </table>
  </div>

  <div class="chart-box" style="margin-top:12px">
    <h3 style="font-size:15px">③ 수록 임계값 (확정)</h3>
    <p style="margin-top:6px"><b>C(위키 등재) OR B(5만+ 팔로워)</b> — 둘 중 하나를 충족하면 명부 수록. 3만~5만 구간은 '워치리스트'로 별도 표시(위키 등재 여부 후속 확인 대상).</p>
  </div>

  <div class="chart-box" style="margin-top:12px">
    <h3 style="font-size:15px">④ 활동성 판정</h3>
    <p style="margin-top:6px">최근 활동(업로드·방송·발매) 기준: <span class="st st-활동중">활동중</span> 3개월 이내 · <span class="st st-확인필요">확인필요</span> 3~12개월(휴면 의심) · <span class="st st-휴면">휴면</span> 12개월+ · <span class="st st-해체종료">해체/종료</span> 공식 발표. MAVE: 같은 사실상 중단 케이스도 자동 탐지합니다.</p>
  </div>

  <div class="chart-box" style="margin-top:12px">
    <h3 style="font-size:15px">⑤ 알려진 한계</h3>
    <ul style="margin:6px 0 0 18px;line-height:1.7">
      <li>그룹 명부는 '위키 등재' 기준이라 미등재 초소형·프리데뷔 팀은 누락될 수 있음</li>
      <li><b>소속 여부 판별 불가</b>: 스트리머 집계 API가 소속사 정보를 제공하지 않아, '수집 스트리머' 목록은 무소속(개인세) 명단이 아니라 <b>소속 미상</b>입니다. 소속이 확인된 솔로는 명부의 <code>버추얼휴먼/솔로</code>로 등재합니다</li>
      <li><b>음반·MV 중심 아티스트는 자동수집 사각지대</b>: 라이브 비중이 낮으면 스트리머 집계에 잡히지 않습니다(예: Hebi.). 이런 아티스트는 <b>명부에 수동 등재</b>하고, 채널 통계만 YouTube API로 자동 갱신합니다</li>
      <li>스트리머 규모는 비공식 단일 API 기반 — 공식 API(치지직/SOOP)로 이전 예정</li>
      <li>일부 소속사·데뷔일은 보도자료 크레딧 기반 '추정'(상세에 명시)</li>
      <li>'확인필요' 항목은 공식 채널 최신 업로드 직접 확인이 필요</li>
    </ul>
  </div>

  <p class="muted" style="margin-top:12px">데이터 수집일 ${D.collected} · 인용 시 출처를 위와 같이 밝혀 주세요.</p>`;
}

/* ── Supabase 인증·권한 ── */
let sb=null, ME=null, MYROLE=null;
if(CFG.supabase && window.supabase){
  try{ sb=window.supabase.createClient(CFG.supabase.url, CFG.supabase.key); }catch(e){ sb=null; }
}
const roleLabel=r=>({admin:'관리자',editor:'에디터',premium:'프리미엄',member:'일반회원'}[r]||'회원');
const isStaff=()=>MYROLE==='editor'||MYROLE==='admin';
const canPremium=()=>MYROLE==='premium'||MYROLE==='editor'||MYROLE==='admin';

async function loadMe(){
  if(!sb){ renderAuthArea(); applyRoleGating(); return; }
  const {data:{user}}=await sb.auth.getUser();
  ME=user||null;
  if(ME){
    const {data}=await sb.from('profiles').select('role,nickname').eq('id',ME.id).maybeSingle();
    MYROLE=data?data.role:'member';
  }else MYROLE=null;
  renderAuthArea(); applyRoleGating();
  if(document.querySelector('.tab.on')&&document.querySelector('.tab.on').dataset.t==='admin') renderAdmin();
}
function renderAuthArea(){
  const el=$('#auth-area'); if(!el) return;
  if(!sb){ el.innerHTML=''; return; }
  if(ME){
    el.innerHTML=`<span class="muted" style="font-size:12px">${esc(ME.email)} <span class="role-pill">${roleLabel(MYROLE)}</span></span> <button class="btn-s" id="btn-logout">로그아웃</button>`;
    $('#btn-logout').onclick=async()=>{ await sb.auth.signOut(); ME=null;MYROLE=null; renderAuthArea(); applyRoleGating(); };
  }else{
    el.innerHTML=`<button class="btn-s" id="btn-login">로그인 / 가입</button>`;
    $('#btn-login').onclick=openAuthModal;
  }
}
function applyRoleGating(){
  const tab=document.querySelector('.tab[data-t="admin"]');
  if(tab) tab.style.display=isStaff()?'':'none';
  if(!isStaff()){ const on=document.querySelector('.tab.on'); if(on&&on.dataset.t==='admin') showTab('dash'); }
}
function openAuthModal(){ $('#am-msg').textContent=''; $('#auth-modal').classList.add('open'); }
function closeAuthModal(){ $('#auth-modal').classList.remove('open'); }
async function doAuth(mode){
  const email=$('#am-email').value.trim(), pw=$('#am-pw').value, msg=$('#am-msg');
  if(!email||!pw){ msg.style.color='#e74c3c'; msg.textContent='이메일과 비밀번호를 입력하세요.'; return; }
  msg.style.color='#7a8194'; msg.textContent='처리 중…';
  const res = mode==='signup' ? await sb.auth.signUp({email,password:pw}) : await sb.auth.signInWithPassword({email,password:pw});
  if(res.error){ msg.style.color='#e74c3c'; msg.textContent=res.error.message; return; }
  if(mode==='signup' && !res.data.session){ msg.style.color='#157a36'; msg.textContent='가입 완료! 이메일 확인이 필요하면 메일을 확인한 뒤 로그인하세요.'; return; }
  msg.style.color='#157a36'; msg.textContent='완료!';
  closeAuthModal(); await loadMe();
}
if($('#am-login')){ $('#am-login').onclick=()=>doAuth('login'); $('#am-signup').onclick=()=>doAuth('signup'); }
if(sb) sb.auth.onAuthStateChange(()=>{ loadMe(); });

/* ── 그룹 제보 폼 (로그인 기반 → Supabase) ── */
(function(){
  const btn=$('#gf-submit'); if(!btn)return;
  btn.onclick=async()=>{
    const val=id=>{const el=$('#gf-'+id);return el?el.value.trim():'';};
    const name=val('name'); const msg=$('#gf-msg');
    if(!name){msg.style.color='#e74c3c';msg.textContent='그룹명은 필수예요.';return;}
    if(!sb){ msg.style.color='#7a5b00'; msg.textContent='제보 시스템이 아직 연결되지 않았어요(공개 사이트에서 이용).'; return; }
    if(!ME){ msg.style.color='#7a5b00'; msg.textContent='제보하려면 먼저 로그인이 필요해요.'; openAuthModal(); return; }
    const row={name,name_en:val('en'),agency:val('agency'),year:val('year'),category:$('#gf-cat').value,status:$('#gf-status').value,channel:val('channel'),note:val('note'),submitter:ME.id};
    msg.style.color='#7a8194'; msg.textContent='접수 중…';
    const {error}=await sb.from('submissions').insert(row);
    if(error){ msg.style.color='#e74c3c'; msg.textContent='접수 실패: '+error.message; return; }
    msg.style.color='#157a36'; msg.textContent='제보가 접수됐어요! 검수자 승인 후 명부에 반영됩니다. 감사합니다.';
    ['name','en','agency','year','channel','note'].forEach(id=>{const e=$('#gf-'+id);if(e)e.value='';});
  };
})();

/* ── 관리자: 제보 검수 + 회원/등급 관리 (Supabase) ── */
async function renderAdmin(){
  const box=$('#admin-body');
  if(!sb){ box.innerHTML='<p class="muted">로그인 시스템이 연결되지 않았어요(공개 사이트에서 이용).</p>'; return; }
  if(!isStaff()){ box.innerHTML='<p class="muted">접근 권한이 없어요. 에디터 이상 등급만 이용할 수 있어요.</p>'; return; }
  box.innerHTML='<p class="muted">불러오는 중…</p>';
  const {data:subs,error}=await sb.from('submissions').select('*').eq('state','pending').order('created_at',{ascending:false});
  if(error){ box.innerHTML='<p class="muted">제보를 불러오지 못했어요: '+esc(error.message)+'</p>'; return; }
  let html=`<h3 style="font-size:15px;margin-bottom:8px">검수 대기 제보 (${subs.length})</h3>`;
  html += subs.length? subs.map(s=>`<div class="chart-box" style="margin-bottom:10px">
      <div style="display:flex;justify-content:space-between;align-items:start;gap:8px"><h3 style="font-size:15px">${esc(s.name)} <span class="muted">${esc(s.name_en||'')}</span></h3><span class="muted">#${s.id} · ${(s.created_at||'').slice(0,10)}</span></div>
      <div class="kv"><div><b>소속사</b><span>${esc(s.agency||'—')}</span></div><div><b>분류·상태</b><span>${esc(s.category||'—')} · ${esc(s.status||'—')}</span></div><div><b>채널</b><span>${s.channel?`<a href="${esc(s.channel)}" target="_blank">링크 ↗</a>`:'—'}</span></div></div>
      ${s.note?`<p class="muted" style="margin-top:6px;line-height:1.5">${esc(s.note)}</p>`:''}
      <div style="display:flex;gap:8px;margin-top:10px"><button class="btn-p sub-approve" data-id="${s.id}">승인</button><button class="btn-s sub-reject" data-id="${s.id}">반려</button></div>
    </div>`).join('') : '<p class="muted">대기 중인 제보가 없어요. 👍</p>';

  if(MYROLE==='admin'){
    const {data:members}=await sb.from('profiles').select('id,email,nickname,role').order('created_at',{ascending:true});
    html+=`<h3 style="font-size:15px;margin:18px 0 8px">회원·등급 관리 (${(members||[]).length})</h3><div class="chart-box">`+
      (members||[]).map(m=>`<div class="mrow"><span>${esc(m.email||m.nickname||m.id.slice(0,8))}</span>
        <select class="role-sel" data-id="${m.id}">
          <option value="member" ${m.role==='member'?'selected':''}>일반회원</option>
          <option value="premium" ${m.role==='premium'?'selected':''}>프리미엄</option>
          <option value="editor" ${m.role==='editor'?'selected':''}>에디터</option>
          <option value="admin" ${m.role==='admin'?'selected':''}>관리자</option>
        </select></div>`).join('')+`</div>`;
  }
  box.innerHTML=html;
  document.querySelectorAll('.sub-approve').forEach(b=>b.onclick=()=>reviewSub(b.dataset.id,'approved'));
  document.querySelectorAll('.sub-reject').forEach(b=>b.onclick=()=>reviewSub(b.dataset.id,'rejected'));
  document.querySelectorAll('.role-sel').forEach(s=>s.onchange=()=>changeRole(s.dataset.id,s.value,s));
}
async function reviewSub(id,state){
  const {error}=await sb.from('submissions').update({state,reviewed_by:ME.id,reviewed_at:new Date().toISOString()}).eq('id',id);
  if(error){ alert('실패: '+error.message); return; }
  await renderAdmin(); if(state==='approved') loadApprovedGroups();
}
async function changeRole(id,role,el){
  const {error}=await sb.from('profiles').update({role}).eq('id',id);
  if(error){ alert('등급 변경 실패: '+error.message); }
  else { el.style.outline='2px solid #27ae60'; setTimeout(()=>el.style.outline='',1200); }
}
const ar=$('#admin-refresh'); if(ar) ar.onclick=renderAdmin;

/* 승인된 제보를 명부에 합쳐 표시 */
async function loadApprovedGroups(){
  if(!sb) return;
  const {data}=await sb.from('submissions').select('*').eq('state','approved');
  if(!data) return;
  let added=false; let maxNo=Math.max(0,...D.groups.map(g=>+g[0]||0));
  data.forEach(s=>{
    if(D.groups.some(g=>g[1]===s.name)) return;
    D.groups.push([++maxNo,s.name,s.name_en||'',s.agency||'(제보)',String(s.year||''),s.category||'스트리밍/MCN형',s.status||'활동중',s.note||'',s.channel||'','제보']);
    added=true;
  });
  if(added){ $('#cnt-g').textContent=D.groups.length; renderG(); }
}

/* ── ⭐ 프리미엄/자료실 (항목별 접근 레벨) ── */
const LV={0:'공개',1:'일반회원',2:'프리미엄',3:'스태프'};
const myLevel=()=>({admin:4,editor:3,premium:2,member:1}[MYROLE]||0);
const lvBadge=l=>`<span class="role-pill">${LV[l]||'프리미엄'} 이상</span>`;
async function renderPremium(){
  const box=$('#premium-body'); const staff=isStaff();
  if(!sb){ box.innerHTML='<div class="chart-box"><h3>⭐ 자료실</h3><p class="muted">로그인 시스템이 연결되지 않았어요(공개 사이트에서 이용).</p></div>'; return; }
  box.innerHTML='<p class="muted">불러오는 중…</p>';
  // RLS가 내 레벨 이상만 내려줌
  const {data,error}=await sb.from('premium_content').select('*').order('created_at',{ascending:false});
  if(error){ box.innerHTML='<p class="muted">불러오기 실패: '+esc(error.message)+' (Supabase에 phase3 마이그레이션을 실행했는지 확인)</p>'; return; }
  let html=`<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px"><h3 style="font-size:16px">⭐ 자료실</h3><span class="role-pill">${roleLabel(MYROLE)||'비회원'} · 열람가능 ${data.length}건</span></div>
    <p class="muted" style="margin-bottom:12px;line-height:1.5">가공 분석·인사이트·데이터셋. 자료마다 접근 등급이 달라요${staff?' (아래 드롭다운으로 항목별 권한 조정).':'. 더 많은 자료는 상위 등급(프리미엄+)에서 열려요.'}</p>`;
  // 스태프: 작성 폼(등급 선택 포함)
  if(staff){
    html+=`<div class="chart-box" style="margin-bottom:12px">
      <h3 style="font-size:14px">새 자료 작성</h3>
      <input type="text" id="pc-title" style="width:100%;margin-top:8px" placeholder="제목">
      <textarea id="pc-body" rows="4" style="width:100%;margin-top:8px;padding:8px 10px;border:1px solid #d7dbe8;border-radius:8px;font-size:13px;font-family:inherit" placeholder="가공 분석·인사이트·데이터 설명"></textarea>
      <div style="display:flex;align-items:center;gap:8px;margin-top:8px">
        <label style="font-size:12px;color:#555e76;font-weight:600">접근등급</label>
        <select id="pc-level"><option value="0">공개</option><option value="1">일반회원+</option><option value="2" selected>프리미엄+</option><option value="3">스태프</option></select>
        <button class="btn-p" id="pc-save">등록</button>
        <span id="pc-msg" class="muted"></span>
      </div>
    </div>`;
  }
  html+= data.length? data.map(c=>`<div class="chart-box" style="margin-bottom:10px">
      <div style="display:flex;justify-content:space-between;align-items:start;gap:8px"><h3 style="font-size:15px">${esc(c.title)}</h3>
        <span style="display:flex;align-items:center;gap:6px">${lvBadge(c.min_level)}<span class="muted">${(c.created_at||'').slice(0,10)}</span></span></div>
      <p style="margin-top:8px;line-height:1.6;white-space:pre-wrap">${esc(c.body||'')}</p>
      ${staff?`<div style="display:flex;align-items:center;gap:8px;margin-top:10px">
        <label style="font-size:12px;color:#555e76;font-weight:600">접근등급</label>
        <select class="pc-lvl" data-id="${c.id}"><option value="0" ${c.min_level===0?'selected':''}>공개</option><option value="1" ${c.min_level===1?'selected':''}>일반회원+</option><option value="2" ${c.min_level===2?'selected':''}>프리미엄+</option><option value="3" ${c.min_level===3?'selected':''}>스태프</option></select>
        <button class="btn-s pc-del" data-id="${c.id}">삭제</button></div>`:''}
    </div>`).join('') : `<div class="chart-box" style="text-align:center;padding:24px"><div style="font-size:30px">🔒</div><p class="muted" style="margin-top:8px;line-height:1.6">열람 가능한 자료가 아직 없어요.${myLevel()<2?'<br>가공 데이터는 <b>프리미엄 등급</b>부터 더 많이 열려요.'+(ME?' 관리자에게 등급을 요청해 주세요.':''):''}</p>${ME?'':'<button class="btn-p" style="margin-top:10px" onclick="openAuthModal()">로그인 / 가입</button>'}</div>`;
  box.innerHTML=html;
  if(staff){
    $('#pc-save').onclick=async()=>{
      const title=$('#pc-title').value.trim(), body=$('#pc-body').value.trim(), min_level=+$('#pc-level').value, msg=$('#pc-msg');
      if(!title){msg.style.color='#e74c3c';msg.textContent='제목을 입력하세요.';return;}
      msg.style.color='#7a8194';msg.textContent='등록 중…';
      const {error}=await sb.from('premium_content').insert({title,body,min_level,created_by:ME.id});
      if(error){msg.style.color='#e74c3c';msg.textContent='실패: '+error.message;return;}
      renderPremium();
    };
    document.querySelectorAll('.pc-lvl').forEach(s=>s.onchange=async()=>{
      const {error}=await sb.from('premium_content').update({min_level:+s.value}).eq('id',s.dataset.id);
      if(error) alert('변경 실패: '+error.message);
      else { s.style.outline='2px solid #27ae60'; setTimeout(()=>s.style.outline='',1200); }
    });
    document.querySelectorAll('.pc-del').forEach(b=>b.onclick=async()=>{
      if(!confirm('삭제할까요?'))return;
      await sb.from('premium_content').delete().eq('id',b.dataset.id); renderPremium();
    });
  }
}

$('#cyear').textContent=new Date().getFullYear();
renderG();renderS();renderTrend();renderA();renderU();renderTL();renderMethod();
routeHash();
window.addEventListener('hashchange',routeHash);
loadMe(); loadApprovedGroups();
</script></body></html>'''

SB_SCRIPT = ('<script src="https://cdn.jsdelivr.net/npm/@supabase/supabase-js@2"></script>'
             if cfg.get('supabase') else '')
html = (HTML.replace('__DATA__', DATA)
            .replace('__TASKID__', 'vtuber-monthly-collect')
            .replace('__SUPABASE_SCRIPT__', SB_SCRIPT))
open(OUT, 'w', encoding='utf-8').write(html)
print('written', OUT, len(html), 'bytes,', len(groups), 'groups,', len(solos), 'solos')
